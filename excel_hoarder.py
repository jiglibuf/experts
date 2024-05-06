import openpyxl
import pandas as pd
from replacements import Replacements

class ExcelHoarder:
    def __init__(self, replacements:Replacements):
        self.replacements = replacements
        self.name = ''
        self.purpose = ''
        self.area = ''
        self.postive_floor = ''
        self.negative_floor = ''
        self.wall_type = ''
        self.foundation_year = ''
        self.kn_zu = ''
        self.type_p = ''
        self.template_path = ''
        # Read the Excel file
        try:
            self.data = pd.read_excel('данные\Расчеты\ЗП НЗ_ЗП ЖЗ_ЗП ОНС.xlsx', dtype= object ,decimal=',', header=0)
            self.template_path_data = pd.read_excel('данные\Способ определения.xlsx')
            # self.register_data = pd.read_excel(r'данные\06-08 Журнал регистрации Разъяснений.xlsx', header=1)
            self.market_analyst= pd.read_excel(r'данные\Анализ рынка.xlsx', header=1)
        except FileNotFoundError:
            print("File not found. Please check the file path.")

    def get_template_path(self):
        if 'Кадастровый номер' in self.template_path_data.columns:
            # match = self.template_path_data.loc[self.template_path_data['Кадастровый номер'] == self.replacements.conditions['<КадастровыйНомер>'],['Модель оценки, использованная при определении КС']]
            match = self.template_path_data.loc[self.template_path_data['Кадастровый номер'] == self.replacements.kn_input_value,['Модель оценки, использованная при определении КС']]
            model = match.iloc[0,0]
            # все большие строки временны в массиве
            if model in ['ЗП НЗ','ЗП ЖЗ', 'ЗП ГСК','''ЗП НЗ 
ДП склад''','''ЗП НЗ 
ДП офис-торг''','''ЗП НЗ 
ДП производство''']:
                self.template_path = 'данные\Шаблоны РЗ\РЗ-НЗ_ЖЗ_ГСК.docx'
            elif model in ['ЗП ОНС']:
                self.template_path = 'данные\Шаблоны РЗ\Ответ РЗ-ОНС.docx'
            else:
                print('Нет шаблона')
            if not match.empty:
                self.replacements.update_conditions('<Модель определения>',model)
                self.replacements.update_template_path(self.template_path)
            else:
                print('Пусто в template_path_data по этому кн')

    def tables_page1(self):
        # Check if 'Кадастровый номер' exists in the data columns
        if 'Кадастровый номер' in self.data.columns:
            # Find rows where 'Кадастровый номер' matches kn_input_value
            matches = self.data.loc[self.data['Кадастровый номер'] == self.replacements.kn_input_value]

            # Check if there are any matches
            if not matches.empty:
                column_counts = {}
                for column_name in matches.columns:
                    # Проверяем, что значение не NaN и имя столбца не содержит "Unnamed"
                    if not column_name.startswith('Unnamed'):
                        # Получаем значение из таблицы и обновляем соответствующее условие
                        value = str(matches.iloc[0][column_name])
                        if value != 'nan':
                            # if value == 'сведения отсутствуют':
                            #     value = '-'
                            # Создаем ключ в нужном формате и обновляем условие
                            if column_name in column_counts:
                                column_counts[column_name] += 1
                                key = f"<{column_name} {column_counts[column_name]}>"
                            else:
                                column_counts[column_name] = 0
                                key = f"<{column_name}>"
                            value = self.process_percents_and_rounding(key, value)
                            self.replacements.update_conditions(key, anchor=value)

                        

            else:
                print(f"No building name found for 'Кадастровый номер': {self.replacements.conditions['<КадастровыйНомер>']}")
        else:
            print("'Кадастровый номер' column not found in the Excel data.")
        try:
            self.replacements.conditions['<Кадастровый номер земельного участка>']
            
        except:
            self.replacements.update_conditions('<Кадастровый номер земельного участка>',anchor='-')

    def process_p_5(self):
        if self.template_path == 'данные\Шаблоны РЗ\РЗ-НЗ_ЖЗ_ГСК.docx':
            if 'Назначение объекта недвижимости ' in self.market_analyst.columns:
                # Find rows where 'Кадастровый номер' matches kn_input_value
                matches = self.market_analyst.loc[self.market_analyst['Назначение объекта недвижимости '] == self.replacements.conditions['<Сегмент+Типовая зона>']]
                # Check if there are any matches
                if not matches.empty:
                    # if self.market_analyst['Назначение объекта недвижимости '] == self.replacements.conditions['Сегмент+Типовая зона']:
                    t = str(matches.iloc[0]['В анализ рынка'])
                    self.replacements.update_conditions('<Решения>', anchor=f'''Удельный показатель кадастровой стоимости объекта недвижимости                                        в размере {self.replacements.conditions['<УПКС 2023>']} руб./кв.м. входит в диапазон рынка недвижимости объектов {t}''')
                else:
                    # self.replacements.update_conditions('<Решения>', anchor=f'''производственного назначения, за исключением передаточных устройств и сооружений Петропавловск-Камчатского городского округа, который находится в границах от 5 175,80 до 156 739,81 руб./кв.м., что ниже среднего значения ХХХ''')
                    
                    self.replacements.update_conditions('<Решения>', anchor=f'''Расчёт кадастровой стоимости соответствует требованиям Методических указаний.''')
        
        if self.replacements.conditions['<Справочник>'] =='УПВС':
            self.replacements.update_conditions('<воспроизводство/замещение>',anchor='воспроизводство')
        if self.replacements.conditions['<Справочник>'] =='КО-ИНВЕСТ':
            self.replacements.update_conditions('<воспроизводство/замещение>',anchor='замещение')
        else:
            pass
        try:
            self.replacements.conditions['<Объем в 5.3>']
        except:
            self.replacements.update_conditions('<Объем в 5.3>',anchor='-')

    def process_percents_and_rounding(self, key, value):
        if key in ['<Кадастровая стоимость 2023>']:
            return value+',00'
        if key in ['<Внешнее устаревание>','<Накопленный износ>','Степень готовности ЕГРН','Степень готовности в п. 5.3']:
            return str(round(float(value)*100, 2))
        elif key in ['<УПКС 2023>','<Затраты на замещение / воспроизводство с учетом ПП и ДКЗ>']:
            return str(round(float(value), 2))
        else:
            return value


    def add_row_to_excel(self):
        # Открываем Excel файл
        wb = openpyxl.load_workbook(filename=r'данные\06-08 Журнал регистрации Разъяснений.xlsx')
        ws = wb.active

        # Получаем номер следующего разъяснения
        last_row_index = ws.max_row
        last_number = int(ws.cell(row=last_row_index, column=3).value.split('/')[-1])  # Получаем номер последнего разъяснения
        next_number = last_number + 1
        next_explanation_number = f"№ РЗ-41/2024/{next_number:06d}"  # Форматируем следующий номер разъяснения
        self.replacements.update_attributes(num_request_value = next_explanation_number)
        # Получаем индекс последней строки
        next_row_index = last_row_index + 1

        # Добавляем данные в каждую ячейку новой строки
        ws.cell(row=next_row_index, column=1, value=next_number)  # Увеличиваем значение последнего номера на 1
        ws.cell(row=next_row_index, column=2, value='')  # Берем значение из первой строки
        ws.cell(row=next_row_index, column=3, value=next_explanation_number)
        ws.cell(row=next_row_index, column=4, value='')  # Пустая дата
        ws.cell(row=next_row_index, column=5, value=self.replacements.kn_input_value)
        ws.cell(row=next_row_index, column=6, value='')  # Пустое значение
        ws.cell(row=next_row_index, column=7, value='')  # Пустое значение
        ws.cell(row=next_row_index, column=8, value='')  # Пустое значение
        ws.cell(row=next_row_index, column=9, value=self.replacements.num_incoming_input_value)
        ws.cell(row=next_row_index, column=10, value=self.replacements.request_date_value)

        # Сохраняем изменения
        wb.save(filename=r'данные\06-08 Журнал регистрации Разъяснений.xlsx')
   
    def update_replacements(self):
        self.get_template_path()
        self.tables_page1()
        self.process_p_5()
        self.add_row_to_excel()
